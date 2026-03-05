"""
Pooja Calendar API: serves filtered data from India_Poojas_USA_Muhurta Excel.
Excludes first sheet and ignores date/muhurta columns; returns Pooja Name, Freq, Cal, Local Language.
"""
from pathlib import Path
from datetime import date, datetime, time, timedelta
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
import re
import httpx
from fastapi import APIRouter, Query
from app.config import settings

router = APIRouter()

# Path to Excel (project root)
EXCEL_PATH = Path(__file__).resolve().parent.parent.parent.parent / "India_Poojas_USA_Muhurta_2026_2030.xlsx"

# First sheet to skip
SKIP_SHEET = "📋 Overview & Legend"

# Header row (1-based) in each state sheet
HEADER_ROW = 4
# Columns to return: 0=Pooja Name, 1=Freq, 2=Cal, 3=Local Language / Calendar Ref
# Column 4 = India Date (IST) - used only for year filter, not returned
KEEP_INDICES = (0, 1, 2, 3)
DATE_COLUMN_INDEX = 4  # for year filter

# Map frontend "Type of Pooja" values to Excel Freq column values
POOJA_TYPE_TO_FREQ = {"daily": "Daily", "monthly": "Monthly", "yearly": "Annual"}

# Localized labels for known frequency values
FREQ_LOCALIZED = {
    "te": {"daily": "రోజువారీ", "monthly": "నెలవారీ", "annual": "వార్షిక"},
}

# Common calendar abbreviations/terms
CAL_LOCALIZED = {
    "te": {
        "solar": "సౌర",
        "lunar": "చాంద్ర",
        "solar/lunar": "సౌర/చాంద్ర",
        "lunisolar": "చాంద్రసౌర",
    },
}

# Rows to suppress from final list
EXCLUDED_POOJA_NAMES = {
    "ప్రతిదినము (pratidinam)",
    "pratidinam",
}

# Known date corrections (IST date) for specific yearly poojas when source sheet is inaccurate.
DATE_OVERRIDES_IST = {
    ("maha shivaratri", 2026): date(2026, 2, 15),
}

_FESTIVAL_HINTS = {
    "maha shivaratri": ["maha_shivratri", "maha_shivaratri"],
    "ugadi / gudi padwa": ["ugadi", "gudi_padwa"],
    "sri rama navami": ["rama_navami", "ram_navami"],
    "sri krishna janmashtami": ["janmashtami", "krishna_janmashtami"],
    "ganesh / vinayaka chaturthi": ["ganesh_chaturthi", "vinayaka_chaturthi"],
    "sri devi navaratri": ["shardiya_navratri", "devi_navaratri", "navaratri"],
    "navaratri": ["shardiya_navratri", "navaratri"],
    "ayudha pooja": ["ayudha_puja", "ayudha_pooja"],
    "saraswati / ayudha pooja": ["ayudha_puja", "saraswati_puja"],
    "diwali lakshmi pooja": ["diwali", "lakshmi_puja", "laxmi_puja"],
    "makara sankranti": ["makar_sankranti"],
    "naga panchami": ["nag_panchami", "naga_panchami"],
    "vasanta navaratri": ["chaitra_navratri", "vasant_navratri"],
    "puthandu / tamil new year": ["puthandu", "tamil_new_year"],
}

_FESTIVAL_KEYS_CACHE: dict[tuple[date, str], set[str]] = {}


def _has_telugu_script(text: str) -> bool:
    return any("\u0c00" <= ch <= "\u0c7f" for ch in (text or ""))


def _has_non_latin_script(text: str) -> bool:
    if not text:
        return False
    # Any alphabetic character outside basic Latin range counts as local script.
    return any(ch.isalpha() and ord(ch) > 127 for ch in text)


def _clean_local_language_name(local_language: str) -> str:
    """
    Keep only short display name from local-language column.
    Removes descriptive notes (usually after '|') and long timing notes.
    """
    text = (local_language or "").strip()
    if not text:
        return ""
    # Common pattern in sheet: "Short Name | Description"
    short = text.split("|", 1)[0].strip()
    if not short:
        return ""
    lower_short = short.lower()
    note_markers = (
        "దినమూ", "ఉదయం", "మధ్యాహ్నం", "సాయంత్రం", "ప్రతి నెల",
        "every day", "continuous", "daily", "times",
    )
    if any(marker in lower_short for marker in note_markers):
        return ""
    return short


def _extract_telugu_title(local_language: str) -> str:
    """
    Local-language column often contains 'title | details'. Use only title part.
    Avoid treating descriptive notes as pooja names.
    """
    text = _clean_local_language_name(local_language)
    if not text or not _has_telugu_script(text):
        return ""
    title = text.strip()
    # Descriptive note pattern (e.g. Sandhyavandanam note) should not become pooja name.
    if (
        ("(" in title and "/" in title) or
        any(token in title for token in ("దినమూ", "ఉదయం", "మధ్యాహ్నం", "సాయంత్రం", "ప్రతి నెల"))
    ):
        return ""
    return title


def _extract_local_title(local_language: str) -> str:
    """
    Generic local-language short title (for non-English state language display).
    """
    text = _clean_local_language_name(local_language)
    if not text:
        return ""
    if not _has_non_latin_script(text):
        return ""
    return text


def _normalize_name(text: str) -> str:
    if not text:
        return ""
    return " ".join(text.strip().lower().split())


def _inject_ganesha_pooja_row(
    rows: list[dict[str, str]],
    pooja_type: str | None,
    language_code: str | None,
) -> list[dict[str, str]]:
    """
    Add a simple 'Ganesha Pooja' row independent of Chaturthi festival rows.
    This row should appear in Daily/Monthly/Yearly filters.
    """
    code = (language_code or "").strip().lower()
    t = (pooja_type or "").strip().lower()

    ganesha_name_by_lang = {
        "te": "గణేశ పూజ",
        "gu": "ગણેશ પૂજા",
        "kn": "ಗಣೇಶ ಪೂಜೆ",
        "ta": "கணேஷ பூஜை",
        "ml": "ഗണേശ പൂജ",
        "mr": "गणेश पूजा",
        "hi": "गणेश पूजा",
        "bn": "গণেশ পূজা",
        "or": "ଗଣେଶ ପୂଜା",
        "pa": "ਗਣੇਸ਼ ਪੂਜਾ",
        "as": "গণেশ পূজা",
    }
    pooja_name = ganesha_name_by_lang.get(code, "Ganesha Pooja")
    local_language = pooja_name

    # Show separate entries by frequency.
    if t == "daily":
        freqs = ["Daily"]
    elif t == "monthly":
        freqs = ["Monthly"]
    elif t == "yearly":
        freqs = ["Annual"]
    else:
        freqs = ["Daily", "Monthly", "Annual"]

    out = list(rows)
    for freq in freqs:
        candidate = {
            "pooja_name": pooja_name,
            "freq": FREQ_LOCALIZED.get("te", {}).get(freq.strip().lower(), freq) if code == "te" else freq,
            "cal": "",
            "local_language": local_language,
            "pooja_date": "",
        }
        key = (
            _normalize_name(candidate["pooja_name"]),
            _normalize_name(candidate["freq"]),
            _normalize_name(candidate["cal"]),
            _normalize_name(candidate["local_language"]),
            _normalize_name(candidate["pooja_date"]),
        )
        exists = False
        for item in out:
            existing_key = (
                _normalize_name(item.get("pooja_name", "")),
                _normalize_name(item.get("freq", "")),
                _normalize_name(item.get("cal", "")),
                _normalize_name(item.get("local_language", "")),
                _normalize_name(item.get("pooja_date", "")),
            )
            if existing_key == key:
                exists = True
                break
        if not exists:
            out.insert(0, candidate)
    return out


def _inject_common_pooja_rows(
    rows: list[dict[str, str]],
    pooja_type: str | None,
    language_code: str | None,
) -> list[dict[str, str]]:
    """
    Inject common poojas that should be available across all states.
    - Purnima Pooja / Satyanarayan Pooja: Monthly + Annual
    - Pradosha: Monthly
    - Ekadasi: Monthly
    """
    code = (language_code or "").strip().lower()
    selected_type = (pooja_type or "").strip().lower()

    name_map = {
        "purnima_satyanarayan_monthly": {
            "en": "Purnima Pooja / Satyanarayan Pooja",
            "te": "పౌర్ణిమ పూజ / సత్యనారాయణ పూజ",
            "hi": "पूर्णिमा पूजा / सत्यनारायण पूजा",
        },
        "purnima_satyanarayan_annual": {
            "en": "Purnima Pooja / Satyanarayan Pooja",
            "te": "పౌర్ణిమ పూజ / సత్యనారాయణ పూజ",
            "hi": "पूर्णिमा पूजा / सत्यनारायण पूजा",
        },
        "pradosha_monthly": {
            "en": "Pradosha Shiva Pooja",
            "te": "ప్రదోష శివ పూజ",
            "hi": "प्रदोष शिव पूजा",
        },
        "ekadasi_monthly": {
            "en": "Ekadasi Shiva Pooja",
            "te": "ఏకాదశి శివ పూజ",
            "hi": "एकादशी शिव पूजा",
        },
    }

    templates: list[dict[str, str]] = []
    if selected_type in ("", "monthly"):
        templates.extend([
            {"key": "purnima_satyanarayan_monthly", "freq": "Monthly"},
            {"key": "pradosha_monthly", "freq": "Monthly"},
            {"key": "ekadasi_monthly", "freq": "Monthly"},
        ])
    if selected_type in ("", "yearly"):
        templates.append({"key": "purnima_satyanarayan_annual", "freq": "Annual"})

    out = list(rows)
    for item in templates:
        display_name = (
            name_map[item["key"]].get(code)
            or name_map[item["key"]].get("en")
            or "Pooja"
        )
        freq_display = (
            FREQ_LOCALIZED.get("te", {}).get(item["freq"].strip().lower(), item["freq"])
            if code == "te"
            else item["freq"]
        )
        candidate = {
            "pooja_name": display_name,
            "freq": freq_display,
            "cal": "",
            "local_language": display_name,
            "pooja_date": "",
        }
        ckey = (
            _normalize_name(candidate["pooja_name"]),
            _normalize_name(candidate["freq"]),
            _normalize_name(candidate["cal"]),
            _normalize_name(candidate["local_language"]),
            _normalize_name(candidate["pooja_date"]),
        )
        exists = False
        for existing in out:
            ekey = (
                _normalize_name(existing.get("pooja_name", "")),
                _normalize_name(existing.get("freq", "")),
                _normalize_name(existing.get("cal", "")),
                _normalize_name(existing.get("local_language", "")),
                _normalize_name(existing.get("pooja_date", "")),
            )
            if ekey == ckey:
                exists = True
                break
        if not exists:
            out.insert(0, candidate)
    return out


def _localize_row(
    pooja_name: str,
    freq: str,
    cal: str,
    local_language: str,
    language_code: str | None,
) -> tuple[str, str, str, str]:
    code = (language_code or "").strip().lower()
    if code == "en":
        return pooja_name, freq, cal, _clean_local_language_name(local_language)

    # Keep canonical pooja name from Excel name column; use Telugu title only when safe.
    pooja_name_out = pooja_name
    # Telugu-specific: stricter extraction to avoid note-like strings.
    if code == "te":
        telugu_title = _extract_telugu_title(local_language)
        if telugu_title:
            pooja_name_out = telugu_title
    else:
        local_title = _extract_local_title(local_language)
        if local_title:
            pooja_name_out = local_title

    freq_out = FREQ_LOCALIZED.get("te", {}).get(freq.strip().lower(), freq) if code == "te" else freq
    cal_out = CAL_LOCALIZED.get("te", {}).get(cal.strip().lower(), cal) if code == "te" else cal
    local_language_out = _clean_local_language_name(local_language)

    return pooja_name_out, freq_out, cal_out, local_language_out


def _parse_excel_date(value) -> date | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    m = re.search(r"(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})", s)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)} {m.group(2)} {m.group(3)}", "%d %b %Y").date()
        except ValueError:
            return None
    return None


def _format_pooja_date_for_country(date_val, country: str | None) -> str:
    dt = _parse_excel_date(date_val)
    if not dt:
        return ""
    c = (country or "").strip().lower()
    if c in ("usa", "us", "united states", "united states of america"):
        try:
            ist_midnight = datetime.combine(dt, time(0, 0), tzinfo=ZoneInfo("Asia/Kolkata"))
            et_date = ist_midnight.astimezone(ZoneInfo("America/New_York")).date()
        except ZoneInfoNotFoundError:
            # Fallback when tz database is unavailable in runtime environment.
            # IST midnight maps to previous calendar day in US Eastern.
            et_date = dt - timedelta(days=1)
        return et_date.strftime("%d %b %Y")
    return dt.strftime("%d %b %Y")


def _apply_date_override(dt: date | None, pooja_name: str) -> date | None:
    if not dt:
        return None
    key = (_normalize_name(pooja_name), dt.year)
    return DATE_OVERRIDES_IST.get(key, dt)


def _country_context(country: str | None) -> tuple[str, str, str, str]:
    c = (country or "").strip().lower()
    if c in ("usa", "us", "united states", "united states of america"):
        return "New York", "40.7128", "-74.0060", "-5"
    if c in ("india", ""):
        return "New Delhi", "28.6139", "77.2090", "5.5"
    # fallback
    return "New Delhi", "28.6139", "77.2090", "5.5"


def _fetch_festival_keys_for_date(target_date: date, country: str | None) -> set[str]:
    ckey = (target_date, (country or "").strip().lower())
    if ckey in _FESTIVAL_KEYS_CACHE:
        return _FESTIVAL_KEYS_CACHE[ckey]
    if not settings.divine_api_key or not settings.divine_access_token:
        _FESTIVAL_KEYS_CACHE[ckey] = set()
        return set()
    place, lat, lon, tzone = _country_context(country)
    try:
        with httpx.Client(timeout=12.0) as client:
            response = client.post(
                "https://astroapi-3.divineapi.com/indian-api/v1/date-specific-festivals",
                headers={"Authorization": f"Bearer {settings.divine_access_token}"},
                data={
                    "api_key": settings.divine_api_key,
                    "year": str(target_date.year),
                    "month": str(target_date.month),
                    "day": str(target_date.day),
                    "Place": place,
                    "lat": lat,
                    "lon": lon,
                    "tzone": tzone,
                },
            )
        if response.status_code != 200:
            _FESTIVAL_KEYS_CACHE[ckey] = set()
            return set()
        payload = response.json() if response.content else {}
        data = payload.get("data") if isinstance(payload, dict) else {}
        keys = set(data.keys()) if isinstance(data, dict) else set()
        _FESTIVAL_KEYS_CACHE[ckey] = keys
        return keys
    except Exception:
        _FESTIVAL_KEYS_CACHE[ckey] = set()
        return set()


def _festival_matches_pooja(keys: set[str], pooja_name: str) -> bool:
    if not keys:
        return False
    normalized = _normalize_name(pooja_name)
    hints = _FESTIVAL_HINTS.get(normalized, [])
    if hints:
        low_keys = [k.lower() for k in keys]
        return any(any(h in k for k in low_keys) for h in hints)
    # fallback weak match by major token overlap
    tokens = [t for t in re.split(r"[^a-z]+", normalized) if len(t) >= 4 and t not in ("pooja", "puja", "sri", "shri")]
    if not tokens:
        return False
    low_keys = [k.lower().replace("_", " ") for k in keys]
    return any(sum(1 for t in tokens if t in k) >= 1 for k in low_keys)


def _resolve_yearly_panchang_date(base_date_val, pooja_name: str, country: str | None) -> str:
    base = _apply_date_override(_parse_excel_date(base_date_val), pooja_name)
    if not base:
        return ""
    # If festival endpoint not configured, keep existing behavior.
    if not settings.divine_api_key or not settings.divine_access_token:
        return _format_pooja_date_for_country(base, country)

    # First trust the base date if festival key matches that day.
    base_keys = _fetch_festival_keys_for_date(base, country)
    if _festival_matches_pooja(base_keys, pooja_name):
        return base.strftime("%d %b %Y")

    # Search nearby dates (tithi can shift by timezone/location)
    for delta in range(1, 16):
        for d in (base - timedelta(days=delta), base + timedelta(days=delta)):
            keys = _fetch_festival_keys_for_date(d, country)
            if _festival_matches_pooja(keys, pooja_name):
                return d.strftime("%d %b %Y")

    # Fallback
    return _format_pooja_date_for_country(base, country)


def _get_state_sheets():
    import openpyxl
    if not EXCEL_PATH.exists():
        return []
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    names = [s for s in wb.sheetnames if s != SKIP_SHEET]
    wb.close()
    return names


def _read_sheet(
    state: str,
    year: int | None,
    pooja_type: str | None,
    language_code: str | None = None,
    country: str | None = None,
):
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    if state not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[state]
    rows = list(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True))  # data only
    wb.close()
    compute_yearly_dates = (pooja_type or "").strip().lower() == "yearly"
    result = []
    for row in rows:
        if not row or len(row) <= max(KEEP_INDICES):
            continue
        pooja_name = (row[0] or "").strip()
        local_language = (row[3] or "").strip() if len(row) > 3 else ""
        if not pooja_name or "───" in pooja_name:
            continue
        if pooja_type and len(row) > 1:
            freq = (row[1] or "").strip()
            expected = POOJA_TYPE_TO_FREQ.get((pooja_type or "").strip().lower(), pooja_type.strip())
            if freq and freq != expected:
                continue
        if year is not None and len(row) > DATE_COLUMN_INDEX:
            date_val = row[DATE_COLUMN_INDEX]
            date_str = str(date_val) if date_val else ""
            if date_val and hasattr(date_val, "year") and getattr(date_val, "year", None) != year:
                continue
            if date_str and str(year) not in date_str and "Every day" not in date_str and "Continuous" not in date_str:
                continue
        freq = (row[1] or "").strip() if len(row) > 1 else ""
        cal = (row[2] or "").strip() if len(row) > 2 else ""
        pooja_name_out, freq_out, cal_out, local_language_out = _localize_row(
            pooja_name=pooja_name,
            freq=freq,
            cal=cal,
            local_language=local_language,
            language_code=language_code,
        )

        # Explicitly hide placeholder/generic "Pratidinam" row from results.
        name_candidates = {
            _normalize_name(pooja_name),
            _normalize_name(local_language),
            _normalize_name(pooja_name_out),
            _normalize_name(local_language_out),
        }
        if any(candidate in EXCLUDED_POOJA_NAMES for candidate in name_candidates):
            continue

        result.append({
            "pooja_name": pooja_name_out,
            "freq": freq_out,
            "cal": cal_out,
            "local_language": local_language_out,
            "pooja_date": _resolve_yearly_panchang_date(
                row[DATE_COLUMN_INDEX] if len(row) > DATE_COLUMN_INDEX else None,
                pooja_name,
                country,
            ) if compute_yearly_dates and freq.strip().lower() == "annual" else "",
        })
    # Excel has repeated lines for some poojas (especially Daily); collapse exact duplicates.
    seen: set[tuple[str, str, str, str, str]] = set()
    deduped: list[dict[str, str]] = []
    for item in result:
        key = (
            (item.get("pooja_name") or "").strip().lower(),
            (item.get("freq") or "").strip().lower(),
            (item.get("cal") or "").strip().lower(),
            (item.get("local_language") or "").strip().lower(),
            (item.get("pooja_date") or "").strip().lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    with_ganesha = _inject_ganesha_pooja_row(deduped, pooja_type, language_code)
    return _inject_common_pooja_rows(with_ganesha, pooja_type, language_code)


@router.get("/states")
async def list_states():
    """Return list of state names (sheet names excluding first)."""
    states = _get_state_sheets()
    return {"states": states}


@router.get("/data")
async def get_calendar_data(
    state: str = Query(..., description="State name (e.g. Andhra Pradesh)"),
    year: int | None = Query(None, description="Filter by year (2026-2030)"),
    type: str | None = Query(None, description="Pooja type: Daily, Monthly, or Yearly"),
    language_code: str | None = Query(None, description="Optional display language code, e.g. te"),
    country: str | None = Query(None, description="Current location country"),
):
    """Return table rows for the selected state/year/type. Excludes date/muhurta columns."""
    if not EXCEL_PATH.exists():
        return {"rows": [], "error": "Excel file not found"}
    rows = _read_sheet(state, year, type, language_code, country)
    return {"rows": rows}
